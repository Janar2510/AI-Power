"""Default routes - web client root and login."""

from werkzeug.wrappers import Response
from werkzeug.utils import redirect

from core.tools import config
from core.modules.assets import get_bundle_urls
from core.http.controller import route
from core.http.request import Request


def _is_debug_assets(request: Request) -> bool:
    """Check if debug=assets mode (individual files, no minification)."""
    if request.args.get("debug") == "assets":
        return True
    return config.get_config().get("debug_assets", False)
from core.http.session import ensure_session_csrf
from core.http.auth import (
    authenticate,
    get_session_uid,
    get_session_db,
    get_session_company_id_from_request,
    get_session_lang_from_request,
    set_session_company_id,
    set_session_lang,
    get_session,
    login_response_with_html,
    logout_response,
    _get_registry,
    user_has_totp_enabled,
    create_totp_pending,
    get_totp_pending,
    clear_totp_pending,
    verify_totp_code,
    generate_totp_secret,
    get_totp_provision_uri,
    save_totp_to_user,
    disable_totp_for_user,
    store_totp_setup,
    get_totp_setup,
    clear_totp_setup,
)
from core.sql_db import db_exists, get_cursor


LOGIN_HTML = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8"/>
<meta name="viewport" content="width=device-width, initial-scale=1"/>
<title>ERP Platform - Login</title>
<link rel="stylesheet" href="/web/static/src/scss/webclient.css"/>
<style>
  body {{ margin: 0; font-family: system-ui, sans-serif; background: #f5f5f5; color: #333; min-height: 100vh; }}
  .login-box {{ max-width: 320px; margin: 4rem auto; padding: 2rem; background: white; border: 1px solid #ddd; border-radius: 8px; box-shadow: 0 2px 8px rgba(0,0,0,0.1); }}
  .login-box h1 {{ margin-top: 0; }}
  .login-box input {{ width: 100%; padding: 0.5rem; margin: 0.5rem 0; box-sizing: border-box; border: 1px solid #ccc; border-radius: 4px; }}
  .login-box button {{ width: 100%; padding: 0.75rem; margin-top: 0.5rem; background: #1a1a2e; color: white; border: none; border-radius: 4px; cursor: pointer; font-size: 1rem; }}
  .login-box .error {{ color: #c00; font-size: 0.9rem; }}
</style>
</head>
<body>
<div class="login-box">
  <h1>ERP Platform</h1>
  <form method="post" action="/web/login">
    <input type="text" name="login" placeholder="Login" required autofocus/>
    <input type="password" name="password" placeholder="Password" required/>
    <input type="hidden" name="db" value="{db}"/>
    <button type="submit">Log in</button>
  </form>
  <p class="error">{error}</p>
  <a href="/web/signup?db={db}">Create an account</a>
</div>
</body>
</html>"""

TOTP_HTML = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8"/>
<meta name="viewport" content="width=device-width, initial-scale=1"/>
<title>ERP Platform - Two-Factor Authentication</title>
<link rel="stylesheet" href="/web/static/src/scss/webclient.css"/>
<style>
  body {{ margin: 0; font-family: system-ui, sans-serif; background: #f5f5f5; color: #333; min-height: 100vh; }}
  .login-box {{ max-width: 320px; margin: 4rem auto; padding: 2rem; background: white; border: 1px solid #ddd; border-radius: 8px; box-shadow: 0 2px 8px rgba(0,0,0,0.1); }}
  .login-box h1 {{ margin-top: 0; }}
  .login-box input {{ width: 100%; padding: 0.5rem; margin: 0.5rem 0; box-sizing: border-box; border: 1px solid #ccc; border-radius: 4px; font-size: 1.2rem; letter-spacing: 0.2em; text-align: center; }}
  .login-box button {{ width: 100%; padding: 0.75rem; margin-top: 0.5rem; background: #1a1a2e; color: white; border: none; border-radius: 4px; cursor: pointer; font-size: 1rem; }}
  .login-box .error {{ color: #c00; font-size: 0.9rem; }}
</style>
</head>
<body>
<div class="login-box">
  <h1>Two-Factor Authentication</h1>
  <p>Enter the 6-digit code from your authenticator app.</p>
  <form method="post" action="/web/login/totp">
    <input type="text" name="code" placeholder="000000" maxlength="6" pattern="[0-9]{{6}}" autocomplete="one-time-code" required autofocus/>
    <input type="hidden" name="db" value="{db}"/>
    <button type="submit">Verify</button>
  </form>
  <p class="error">{error}</p>
  <a href="/web/login?db={db}">Back to login</a>
</div>
</body>
</html>"""


@route("/health", auth="public", methods=["GET"])
def health(request):
    """Health check endpoint for load balancers and monitoring (Phase 99)."""
    import json
    db = request.args.get("db", config.get_config().get("db_name", "erp"))
    db_ok = db_exists(db)
    try:
        body = json.dumps({"status": "ok", "db": db_ok})
    except Exception:
        body = '{"status":"ok","db":false}'
    return Response(body, content_type="application/json")


def _try_init_database(db: str) -> tuple[bool, str]:
    """Try to create and initialize database. Returns (success, message)."""
    try:
        from core.sql_db import create_database, get_cursor
        from core.db import init_schema
        from core.db.init_data import load_default_data, assign_admin_groups
        from core.orm import Registry, Environment
        from core.orm.models import ModelBase
        from core.modules import clear_loaded_addon_modules, load_module_graph
        from core.http.auth import hash_password
        if db_exists(db):
            return True, "Database already exists."
        create_database(db)
        registry = Registry(db)
        ModelBase._registry = registry
        clear_loaded_addon_modules()
        load_module_graph()
        with get_cursor(db) as cr:
            init_schema(cr, registry)
            env = Environment(registry, cr=cr, uid=1)
            registry.set_env(env)
            load_default_data(env)
            User = env.get("res.users")
            if User and not User.search([("login", "=", "admin")]):
                User.create({
                    "login": "admin",
                    "password": hash_password("admin"),
                    "name": "Administrator",
                })
            assign_admin_groups(env)
        return True, "Database initialized. Log in with admin / admin."
    except Exception as e:
        return False, str(e)


@route("/web/login", auth="public", methods=["GET", "POST"])
def login(request):
    """Login page - GET shows form, POST authenticates."""
    if request.method == "POST":
        login_name = request.form.get("login", "").strip()
        password = request.form.get("password", "")
        db = request.form.get("db", "erp")
        if not db_exists(db):
            ok, msg = _try_init_database(db)
            if ok:
                return Response(
                    LOGIN_HTML.format(db=db, error=msg or "Database initialized. Log in with admin / admin."),
                    content_type="text/html; charset=utf-8",
                )
            return Response(
                LOGIN_HTML.format(
                    db=db,
                    error=f"Database {db} does not exist. "
                    f"<a href='/web/login?db={db}&init=1'>Click here to initialize</a> or run: erp-bin db init -d {db}. Error: {msg}",
                ),
                content_type="text/html; charset=utf-8",
                status=400,
            )
        uid = authenticate(login_name, password, db)
        if uid:
            if user_has_totp_enabled(uid, db):
                token = create_totp_pending(uid, db)
                resp = redirect(f"/web/login/totp?db={db}")
                resp.set_cookie("erp_totp_pending", token, httponly=True, samesite="Lax", path="/", max_age=300)
                return resp
            return login_response_with_html(
                uid, db, _webclient_html(debug_assets=_is_debug_assets(request))
            )
        return Response(
            LOGIN_HTML.format(db=db, error="Invalid login or password."),
            content_type="text/html; charset=utf-8",
            status=401,
        )
    db = request.args.get("db", "erp")
    error_msg = ""
    if request.args.get("init") == "1":
        ok, msg = _try_init_database(db)
        error_msg = msg if ok else f"Init failed: {msg}"
    return Response(
        LOGIN_HTML.format(db=db, error=error_msg),
        content_type="text/html; charset=utf-8",
    )


@route("/web/login/totp", auth="public", methods=["GET", "POST"])
def login_totp(request):
    """TOTP verification - GET shows form, POST verifies code and completes login (Phase 125)."""
    token = request.cookies.get("erp_totp_pending")
    db = request.form.get("db", request.args.get("db", "erp")) if request.method == "POST" else request.args.get("db", "erp")
    pending = get_totp_pending(token) if token else None
    if not pending:
        resp = redirect(f"/web/login?db={db}")
        resp.delete_cookie("erp_totp_pending", path="/")
        return resp
    uid, db_from_pending = pending
    if request.method == "POST":
        db = request.form.get("db", db_from_pending)
        code = (request.form.get("code") or "").strip()
        if verify_totp_code(uid, db, code):
            clear_totp_pending(token)
            resp = Response(
                _webclient_html(debug_assets=_is_debug_assets(request)),
                content_type="text/html; charset=utf-8",
            )
            from core.http.auth import _get_user_company_id
            from core.http.session import create_session
            company_id = _get_user_company_id(uid, db)
            sid = create_session(uid, db, company_id)
            resp.set_cookie("erp_session", sid, httponly=True, samesite="Lax", path="/")
            resp.delete_cookie("erp_totp_pending", path="/")
            return resp
        return Response(
            TOTP_HTML.format(db=db, error="Invalid code. Please try again."),
            content_type="text/html; charset=utf-8",
            status=401,
        )
    return Response(
        TOTP_HTML.format(db=db_from_pending, error=""),
        content_type="text/html; charset=utf-8",
    )


@route("/web/logout", auth="public", methods=["GET"])
def logout(request):
    """Logout - clear session."""
    return logout_response()


SIGNUP_HTML = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8"/>
<meta name="viewport" content="width=device-width, initial-scale=1"/>
<title>ERP Platform - Sign up</title>
<link rel="stylesheet" href="/web/static/src/scss/webclient.css"/>
<style>
  body {{ margin: 0; font-family: system-ui, sans-serif; background: #f5f5f5; color: #333; min-height: 100vh; }}
  .signup-box {{ max-width: 320px; margin: 4rem auto; padding: 2rem; background: white; border: 1px solid #ddd; border-radius: 8px; box-shadow: 0 2px 8px rgba(0,0,0,0.1); }}
  .signup-box h1 {{ margin-top: 0; }}
  .signup-box input {{ width: 100%; padding: 0.5rem; margin: 0.5rem 0; box-sizing: border-box; border: 1px solid #ccc; border-radius: 4px; }}
  .signup-box button {{ width: 100%; padding: 0.75rem; margin-top: 0.5rem; background: #1a1a2e; color: white; border: none; border-radius: 4px; cursor: pointer; font-size: 1rem; }}
  .signup-box .error {{ color: #c00; font-size: 0.9rem; }}
  .signup-box .success {{ color: #080; font-size: 0.9rem; }}
  .signup-box a {{ margin-top: 1rem; display: block; text-align: center; }}
</style>
</head>
<body>
<div class="signup-box">
  <h1>Sign up</h1>
  <form method="post" action="/web/signup">
    <input type="text" name="name" placeholder="Full name" required/>
    <input type="hidden" name="db" value="{db}"/>
    <input type="email" name="email" placeholder="Email" required/>
    <input type="text" name="login" placeholder="Login (username)" required/>
    <input type="password" name="password" placeholder="Password" required/>
    <button type="submit">Create account</button>
  </form>
  <p class="error">{error}</p>
  <p class="success">{success}</p>
  <a href="/web/login">Already have an account? Log in</a>
</div>
</body>
</html>"""


@route("/web/signup", auth="public", methods=["GET", "POST"])
def signup(request):
    """Signup page - GET shows form, POST creates portal user (Phase 98)."""
    db = request.form.get("db", request.args.get("db", "erp")) if request.method == "POST" else request.args.get("db", "erp")
    if request.method == "POST":
        name = (request.form.get("name") or "").strip()
        email = (request.form.get("email") or "").strip()
        login_name = (request.form.get("login") or "").strip()
        password = request.form.get("password", "")
        if not name or not login_name or not password:
            return Response(
                SIGNUP_HTML.format(db=db, error="Name, login and password are required.", success=""),
                content_type="text/html; charset=utf-8",
                status=400,
            )
        if not db_exists(db):
            return Response(
                SIGNUP_HTML.format(db=db, error=f"Database {db} does not exist.", success=""),
                content_type="text/html; charset=utf-8",
                status=400,
            )
        try:
            with get_cursor(db) as cr:
                from core.orm import Environment
                registry = _get_registry(db)
                env = Environment(registry, cr=cr, uid=1)
                registry.set_env(env)
                User = env.get("res.users")
                uid = User._create_portal_user(env, name=name, login=login_name, password=password, email=email or None)
                if uid:
                    return Response(
                        SIGNUP_HTML.format(db=db, error="", success="Account created. You can now log in."),
                        content_type="text/html; charset=utf-8",
                    )
        except Exception as e:
            pass
        return Response(
            SIGNUP_HTML.format(db=db, error="Signup failed. Login may already exist.", success=""),
            content_type="text/html; charset=utf-8",
            status=400,
        )
    return Response(
        SIGNUP_HTML.format(db=db, error="", success=""),
        content_type="text/html; charset=utf-8",
    )


@route("/web/load_views", auth="user", methods=["GET"])
def load_views(request):
    """Return views, actions, menus from DB (persistent) when session exists, else XML."""
    uid = get_session_uid(request)
    if uid is None:
        return Response('{"error": "unauthorized"}', status=401, content_type="application/json")
    import json
    from core.orm import Environment
    from core.data.views_registry import load_views_registry, load_views_registry_from_db
    from core.db.init_data import load_default_data

    db = get_session_db(request)
    try:
        registry_obj = _get_registry(db)
        with get_cursor(db) as cr:
            env = Environment(registry_obj, cr=cr, uid=uid)
            registry_obj.set_env(env)
            registry = load_views_registry_from_db(env)
            # Phase 170: auto-upgrade when DB menus are empty or stale
            xml_reg = load_views_registry()
            xml_menu_count = len(xml_reg.get("menus", []))
            registry_menus = registry.get("menus", [])
            if xml_menu_count > 0 and (
                len(registry_menus) == 0 or len(registry_menus) < xml_menu_count * 0.5
            ):
                load_default_data(env)
                cr.connection.commit()
                registry = load_views_registry_from_db(env)
            fields_meta = {}
            for model_name in list(registry.get("views", {}).keys()):
                ModelCls = env.get(model_name)
                if ModelCls and hasattr(ModelCls, "fields_get"):
                    try:
                        fields_meta[model_name] = ModelCls.fields_get()
                    except Exception:
                        try:
                            cr.connection.rollback()
                        except Exception:
                            pass
            registry["fields_meta"] = fields_meta
    except Exception:
        registry = load_views_registry()

    def _json_safe(obj):
        """Recursively replace non-JSON-serializable values (e.g. callables)."""
        if callable(obj):
            return None
        if isinstance(obj, dict):
            return {k: _json_safe(v) for k, v in obj.items()}
        if isinstance(obj, (list, tuple)):
            return [_json_safe(x) for x in obj]
        return obj

    return Response(
        json.dumps(_json_safe(registry)),
        content_type="application/json",
    )


def _parse_import_file(file_storage):
    """Parse CSV or XLSX file. Returns (headers, all_rows) or (None, None) on error. Phase 179."""
    if not file_storage or not file_storage.filename:
        return None, None
    fn = (file_storage.filename or "").lower()
    try:
        if fn.endswith(".xlsx"):
            from openpyxl import load_workbook
            wb = load_workbook(file_storage, read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            wb.close()
            if not rows:
                return None, None
            headers = [str(c or "").strip() for c in rows[0]]
            data_rows = [[str(c or "").strip() if c is not None else "" for c in row] for row in rows[1:]]
            return headers, data_rows
        else:
            import csv
            import io
            content = file_storage.read()
            if isinstance(content, bytes):
                content = content.decode("utf-8", errors="replace")
            reader = csv.reader(io.StringIO(content))
            rows = list(reader)
            if not rows:
                return None, None
            headers = [str(c or "").strip() for c in rows[0]]
            data_rows = [[str(c or "").strip() for c in row] for row in rows[1:]]
            return headers, data_rows
    except Exception:
        return None, None


@route("/web/import/preview", auth="user", methods=["POST"])
def import_preview(request):
    """Phase 179: Parse CSV or XLSX, return column headers + first 5 rows."""
    uid = get_session_uid(request)
    if uid is None:
        return Response('{"error": "unauthorized"}', status=401, content_type="application/json")
    import json
    f = request.files.get("file")
    headers, rows = _parse_import_file(f)
    if headers is None:
        return Response(json.dumps({"error": "Could not parse file. Use CSV or XLSX."}), status=400, content_type="application/json")
    preview = rows[:5] if rows else []
    return Response(json.dumps({"headers": headers, "rows": preview, "total_rows": len(rows) if rows else 0}), content_type="application/json")


@route("/web/import/execute", auth="user", methods=["POST"])
def import_execute(request):
    """Phase 179: Import file with column mapping. Returns {success, errors}."""
    uid = get_session_uid(request)
    if uid is None:
        return Response('{"error": "unauthorized"}', status=401, content_type="application/json")
    import json
    f = request.files.get("file")
    model = (request.form.get("model") or "").strip()
    mapping_str = request.form.get("mapping") or "{}"
    if not model or not f:
        return Response(json.dumps({"error": "model and file required"}), status=400, content_type="application/json")
    try:
        mapping = json.loads(mapping_str)
    except Exception:
        mapping = {}
    headers, rows = _parse_import_file(f)
    if headers is None or not rows:
        return Response(json.dumps({"error": "Could not parse file"}), status=400, content_type="application/json")
    db = get_session_db(request)
    registry_obj = _get_registry(db)
    with get_cursor(db) as cr:
        from core.orm import Environment
        env = Environment(registry_obj, cr=cr, uid=uid)
        registry_obj.set_env(env)
        ModelCls = env.get(model)
        if not ModelCls:
            return Response(json.dumps({"error": "model not found"}), status=404, content_type="application/json")
        col_to_field = {}
        for i in range(len(headers)):
            fn = mapping.get(str(i))
            if fn and fn != "id":
                col_to_field[i] = fn
        fields = list(dict.fromkeys(col_to_field.values()))
        if not fields:
            return Response(json.dumps({"error": "map at least one column"}), status=400, content_type="application/json")
        ordered_rows = []
        for row in rows:
            vals = []
            for fn in fields:
                found = ""
                for idx, f in col_to_field.items():
                    if f == fn and idx < len(row):
                        found = row[idx] if row[idx] is not None else ""
                        break
                vals.append(str(found).strip() if found else "")
            ordered_rows.append(vals)
        try:
            res = ModelCls.import_data(fields, ordered_rows)
            return Response(json.dumps({
                "success": (res.get("created") or 0) + (res.get("updated") or 0),
                "created": res.get("created", 0),
                "updated": res.get("updated", 0),
                "errors": res.get("errors", []),
            }), content_type="application/json")
        except Exception as e:
            return Response(json.dumps({"error": str(e), "success": 0, "errors": []}), status=500, content_type="application/json")


@route("/web/import/bank_statement", auth="user", methods=["POST"])
def import_bank_statement(request):
    """Phase 193: Import bank statement from CSV. Expects columns: date, amount, name; optional: partner."""
    uid = get_session_uid(request)
    if uid is None:
        return Response('{"error": "unauthorized"}', status=401, content_type="application/json")
    import json
    f = request.files.get("file")
    statement_id_str = (request.form.get("statement_id") or "").strip()
    journal_id_str = (request.form.get("journal_id") or "").strip()
    date_str = (request.form.get("date") or "").strip()
    if not f:
        return Response(json.dumps({"error": "file required"}), status=400, content_type="application/json")
    headers, rows = _parse_import_file(f)
    if headers is None or not rows:
        return Response(json.dumps({"error": "Could not parse file. Use CSV."}), status=400, content_type="application/json")
    db = get_session_db(request)
    registry_obj = _get_registry(db)
    with get_cursor(db) as cr:
        from core.orm import Environment
        env = Environment(registry_obj, cr=cr, uid=uid)
        registry_obj.set_env(env)
        Statement = env.get("account.bank.statement")
        Line = env.get("account.bank.statement.line")
        Journal = env.get("account.journal")
        Partner = env.get("res.partner")
        if not Statement or not Line:
            return Response(json.dumps({"error": "Bank statement models not found"}), status=500, content_type="application/json")
        h = [str(x).lower() for x in headers]
        date_col = next((i for i, x in enumerate(h) if x in ("date", "transaction date")), 0)
        amount_col = next((i for i, x in enumerate(h) if x in ("amount", "credit", "debit", "value")), 1)
        name_col = next((i for i, x in enumerate(h) if x in ("name", "description", "memo", "reference")), 2)
        partner_col = next((i for i, x in enumerate(h) if x in ("partner", "counterparty", "payee")), None)
        sid = int(statement_id_str) if statement_id_str and statement_id_str.isdigit() else None
        if sid:
            st = Statement.browse(sid)
            if not st.ids:
                return Response(json.dumps({"error": "Statement not found"}), status=404, content_type="application/json")
        else:
            if not journal_id_str or not date_str:
                return Response(json.dumps({"error": "journal_id and date required when creating new statement"}), status=400, content_type="application/json")
            try:
                jid = int(journal_id_str)
            except ValueError:
                return Response(json.dumps({"error": "invalid journal_id"}), status=400, content_type="application/json")
            if not Journal.browse(jid).ids:
                return Response(json.dumps({"error": "Journal not found"}), status=404, content_type="application/json")
            st = Statement.create({"journal_id": jid, "date": date_str[:10], "balance_start": 0.0})
            sid = st.ids[0]
        created = 0
        for row in rows:
            if len(row) <= max(date_col, amount_col, name_col):
                continue
            date_val = str(row[date_col] or "")[:10] if date_col < len(row) else date_str[:10]
            try:
                amt = float(str(row[amount_col] or "0").replace(",", "."))
            except (ValueError, TypeError):
                continue
            name_val = str(row[name_col] or "Import")[:200] if name_col < len(row) else "Import"
            partner_id = None
            if partner_col is not None and partner_col < len(row) and row[partner_col]:
                pname = str(row[partner_col]).strip()
                if pname:
                    partners = Partner.search([("name", "ilike", pname)], limit=1)
                    if partners.ids:
                        partner_id = partners.ids[0]
            Line.create({
                "statement_id": sid,
                "name": name_val,
                "date": date_val,
                "amount": amt,
                "partner_id": partner_id,
            })
            created += 1
        return Response(json.dumps({"success": created, "line_count": created, "statement_id": sid}), content_type="application/json")


@route("/web/attachment/upload", auth="user", methods=["POST"])
def attachment_upload(request):
    """Phase 184: Upload file, create ir.attachment. Returns {id, name, mimetype, url}."""
    uid = get_session_uid(request)
    if uid is None:
        return Response('{"error": "unauthorized"}', status=401, content_type="application/json")
    import json
    f = request.files.get("file")
    res_model = (request.form.get("res_model") or "").strip()
    res_id_str = (request.form.get("res_id") or "").strip()
    if not f or not res_model or not res_id_str:
        return Response(json.dumps({"error": "file, res_model, res_id required"}), status=400, content_type="application/json")
    try:
        res_id = int(res_id_str)
    except ValueError:
        return Response(json.dumps({"error": "invalid res_id"}), status=400, content_type="application/json")
    db = get_session_db(request)
    registry_obj = _get_registry(db)
    with get_cursor(db) as cr:
        from core.orm import Environment
        env = Environment(registry_obj, cr=cr, uid=uid)
        registry_obj.set_env(env)
        Attachment = env.get("ir.attachment")
        if not Attachment:
            return Response(json.dumps({"error": "ir.attachment not found"}), status=500, content_type="application/json")
        name = (f.filename or "attachment").strip() or "attachment"
        data = f.read()
        import base64
        b64 = base64.b64encode(data).decode("ascii") if data else ""
        mimetype = getattr(f, "content_type", None) or "application/octet-stream"
        att = Attachment.create({
            "name": name,
            "res_model": res_model,
            "res_id": res_id,
            "datas": b64,
            "mimetype": mimetype,
        })
        aid = att.ids[0] if att.ids else att.id
        return Response(json.dumps({
            "id": aid,
            "name": name,
            "mimetype": mimetype,
            "url": f"/web/attachment/download/{aid}",
        }), content_type="application/json")


def _attachment_download_view(request, att_id: int):
    """Phase 184/212: Serve attachment file (called from application for /web/attachment/download/<id>)."""
    uid = get_session_uid(request)
    if uid is None:
        return Response("Unauthorized", status=401)
    db = get_session_db(request)
    registry_obj = _get_registry(db)
    with get_cursor(db) as cr:
        from core.orm import Environment
        env = Environment(registry_obj, cr=cr, uid=uid)
        registry_obj.set_env(env)
        Attachment = env.get("ir.attachment")
        if not Attachment:
            return Response("Not found", status=404)
        atts = Attachment.browse([att_id]).read(["name", "datas", "mimetype"])
        if not atts or not atts[0].get("datas"):
            return Response("Not found", status=404)
        import base64
        data = base64.b64decode(atts[0]["datas"])
        name = atts[0].get("name", "download")
        mimetype = atts[0].get("mimetype") or "application/octet-stream"
        from werkzeug.utils import secure_filename
        safe_name = secure_filename(name) or "download"
        resp = Response(data, mimetype=mimetype)
        resp.headers["Content-Disposition"] = f'attachment; filename="{safe_name}"'
        return resp


@route("/web/binary/upload", auth="user", methods=["POST"])
def binary_upload(request):
    """Phase 212: Odoo-style binary upload. Creates ir.attachment. Optional res_model, res_id."""
    uid = get_session_uid(request)
    if uid is None:
        return Response('{"error": "unauthorized"}', status=401, content_type="application/json")
    import json
    f = request.files.get("ufile") or request.files.get("file")
    res_model = (request.form.get("res_model") or "").strip()
    res_id_str = (request.form.get("res_id") or "").strip()
    if not f:
        return Response(json.dumps({"error": "file required (ufile or file)"}), status=400, content_type="application/json")
    res_id = int(res_id_str) if res_id_str and res_id_str.isdigit() else None
    db = get_session_db(request)
    registry_obj = _get_registry(db)
    with get_cursor(db) as cr:
        from core.orm import Environment
        env = Environment(registry_obj, cr=cr, uid=uid)
        registry_obj.set_env(env)
        Attachment = env.get("ir.attachment")
        if not Attachment:
            return Response(json.dumps({"error": "ir.attachment not found"}), status=500, content_type="application/json")
        name = (f.filename or "attachment").strip() or "attachment"
        data = f.read()
        import base64
        b64 = base64.b64encode(data).decode("ascii") if data else ""
        vals = {"name": name, "datas": b64}
        if res_model:
            vals["res_model"] = res_model
        if res_id is not None:
            vals["res_id"] = res_id
        att = Attachment.create(vals)
        aid = att.ids[0] if att.ids else att.id
        return Response(json.dumps({
            "id": aid,
            "name": name,
            "mimetype": getattr(f, "content_type", None) or "application/octet-stream",
            "url": f"/web/attachment/download/{aid}",
        }), content_type="application/json")


def _content_download_view(request, model: str, rec_id: int, field: str, filename: str):
    """Phase 212: Serve binary field from any model. GET /web/content/<model>/<id>/<field>/<filename>."""
    uid = get_session_uid(request)
    if uid is None:
        return Response("Unauthorized", status=401)
    db = get_session_db(request)
    registry_obj = _get_registry(db)
    with get_cursor(db) as cr:
        from core.orm import Environment
        env = Environment(registry_obj, cr=cr, uid=uid)
        registry_obj.set_env(env)
        ModelCls = env.get(model)
        if not ModelCls:
            return Response("Not found", status=404)
        atts = ModelCls.browse([rec_id]).read([field])
        if not atts or not atts[0].get(field):
            return Response("Not found", status=404)
        import base64
        data = base64.b64decode(atts[0][field])
        from werkzeug.utils import secure_filename
        safe_name = secure_filename(filename or "download") or "download"
        resp = Response(data, mimetype="application/octet-stream")
        resp.headers["Content-Disposition"] = f'attachment; filename="{safe_name}"'
        return resp


@route("/web/export/xlsx", auth="user", methods=["POST"])
def export_xlsx(request):
    """Phase 174: Server-side Excel export. POST {model, fields, domain} -> .xlsx binary."""
    uid = get_session_uid(request)
    if uid is None:
        return Response('{"error": "unauthorized"}', status=401, content_type="application/json")
    import json
    from core.orm import Environment
    from io import BytesIO
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except ImportError:
        return Response('{"error": "openpyxl not installed"}', status=500, content_type="application/json")
    try:
        data = request.get_json() or {}
        model = data.get("model", "")
        fields = data.get("fields", ["id", "name"])
        domain = data.get("domain", [])
        if not model:
            return Response('{"error": "model required"}', status=400, content_type="application/json")
        db = get_session_db(request)
        registry_obj = _get_registry(db)
        with get_cursor(db) as cr:
            env = Environment(registry_obj, cr=cr, uid=uid)
            registry_obj.set_env(env)
            ModelCls = env.get(model)
            if not ModelCls:
                return Response('{"error": "model not found"}', status=404, content_type="application/json")
            rows = ModelCls.search_read(domain, fields, limit=10000)
        wb = Workbook()
        ws = wb.active
        ws.title = model.replace(".", "_")[:31]
        for col, f in enumerate(fields, 1):
            ws.cell(row=1, column=col, value=str(f))
        for row_idx, r in enumerate(rows, 2):
            for col_idx, f in enumerate(fields, 1):
                val = r.get(f)
                if isinstance(val, (list, tuple)) and len(val) >= 2:
                    val = val[1]
                elif val is None:
                    val = ""
                ws.cell(row=row_idx, column=col_idx, value=val)
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return Response(
            buf.getvalue(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={model.replace('.', '_')}_export.xlsx"},
        )
    except Exception as e:
        return Response(json.dumps({"error": str(e)}), status=500, content_type="application/json")


@route("/web/session/get_session_info", auth="public", methods=["POST"])
def get_session_info(request):
    """Return session info if valid. Used by frontend to verify cookie is working."""
    uid = get_session_uid(request)
    db = get_session_db(request)
    if uid is None:
        return Response(
            '{"error": "no session"}',
            content_type="application/json",
            status=401,
        )
    import json
    sid = request.cookies.get("erp_session")
    csrf_token = ensure_session_csrf(sid) if sid else None
    result = {"uid": uid, "db": db, "lang": get_session_lang_from_request(request), "csrf_token": csrf_token}
    # Phase 90: user_companies for company switcher
    try:
        registry_obj = _get_registry(db)
        with get_cursor(db) as cr:
            from core.orm import Environment
            env = Environment(registry_obj, cr=cr, uid=uid)
            registry_obj.set_env(env)
            User = env["res.users"]
            Company = env.get("res.company")
            rows = User.read_ids([uid], ["company_id", "company_ids"])
            company_ids = []
            if rows:
                company_ids = rows[0].get("company_ids") or []
                if not company_ids and rows[0].get("company_id"):
                    company_ids = [rows[0]["company_id"]]
            current_id = get_session_company_id_from_request(request)
            if not current_id and company_ids:
                current_id = company_ids[0]
            allowed = []
            if Company and company_ids:
                comp_rows = Company.read_ids(company_ids, ["id", "name"])
                for r in comp_rows:
                    allowed.append({"id": r["id"], "name": r.get("name", "")})
            result["user_companies"] = {
                "current_company": next((a for a in allowed if a["id"] == current_id), allowed[0] if allowed else None),
                "allowed_companies": allowed,
            }
    except Exception:
        result["user_companies"] = {"current_company": None, "allowed_companies": []}
    try:
        registry_obj = _get_registry(db)
        with get_cursor(db) as cr:
            from core.orm import Environment
            env = Environment(registry_obj, cr=cr, uid=uid)
            Lang = env.get("res.lang")
        if Lang:
            rows = Lang.search_read([["active", "=", True]], ["code", "name"], order="name")
            result["user_langs"] = [{"code": r.get("code"), "name": r.get("name")} for r in (rows or [])]
        else:
            result["user_langs"] = [{"code": "en_US", "name": "English"}]
    except Exception:
        result["user_langs"] = [{"code": "en_US", "name": "English"}]
    # Phase 101: user groups for portal detection
    try:
        from core.orm.security import get_user_groups
        registry_obj = _get_registry(db)
        result["groups"] = list(get_user_groups(registry_obj, db, uid))
    except Exception:
        result["groups"] = []
    return Response(
        json.dumps(result),
        content_type="application/json",
    )


@route("/web/translations", auth="user", methods=["GET"])
def get_translations(request):
    """Return translation catalog for session language (Phase 94)."""
    uid = get_session_uid(request)
    if uid is None:
        return Response('{"error": "unauthorized"}', status=401, content_type="application/json")
    import json
    lang = request.args.get("lang") or get_session_lang_from_request(request)
    try:
        registry_obj = _get_registry(get_session_db(request))
        with get_cursor(get_session_db(request)) as cr:
            from core.orm import Environment
            env = Environment(registry_obj, cr=cr, uid=uid)
            registry_obj.set_env(env)
            Translation = env.get("ir.translation")
            catalog = {}
            if Translation:
                rows = Translation.search_read([["lang", "=", lang]], ["src", "value"])
                for r in rows:
                    src = (r.get("src") or "").strip()
                    val = (r.get("value") or "").strip()
                    if src and val:
                        catalog[src] = val
        return Response(json.dumps(catalog), content_type="application/json")
    except Exception:
        return Response("{}", content_type="application/json")


@route("/web/session/set_lang", auth="user", methods=["POST"])
def set_lang(request):
    """Set session language. Requires lang in JSON body."""
    uid = get_session_uid(request)
    if uid is None:
        return Response('{"error": "unauthorized"}', status=401, content_type="application/json")
    try:
        data = request.get_json(force=True, silent=True) or {}
        lang = (data.get("lang") or "").strip() or "en_US"
        sid = request.cookies.get("erp_session")
        if not sid or not get_session(sid):
            return Response('{"error": "no session"}', status=401, content_type="application/json")
        set_session_lang(sid, lang)
        return Response('{"ok": true}', content_type="application/json")
    except Exception:
        return Response('{"error": "invalid"}', status=400, content_type="application/json")


@route("/web/session/set_current_company", auth="user", methods=["POST"])
def set_current_company(request):
    """Set current company in session. Requires company_id in JSON body."""
    uid = get_session_uid(request)
    if uid is None:
        return Response('{"error": "unauthorized"}', status=401, content_type="application/json")
    try:
        data = request.get_json(force=True, silent=True) or {}
        company_id = data.get("company_id")
        if company_id is None:
            return Response('{"error": "company_id required"}', status=400, content_type="application/json")
        company_id = int(company_id)
        sid = request.cookies.get("erp_session")
        if not sid or not get_session(sid):
            return Response('{"error": "no session"}', status=401, content_type="application/json")
        db = get_session_db(request)
        registry_obj = _get_registry(db)
        with get_cursor(db) as cr:
            from core.orm import Environment
            env = Environment(registry_obj, cr=cr, uid=uid)
            User = env["res.users"]
            rows = User.read_ids([uid], ["company_ids"])
            allowed = rows[0].get("company_ids") or [] if rows else []
            if company_id not in allowed:
                return Response('{"error": "company not allowed"}', status=403, content_type="application/json")
        set_session_company_id(sid, company_id)
        return Response('{"ok": true}', content_type="application/json")
    except (ValueError, TypeError):
        return Response('{"error": "invalid company_id"}', status=400, content_type="application/json")


def _webclient_html(debug_assets: bool = False) -> str:
    """Web client shell HTML. Use debug_assets=True for individual files (no minification)."""
    if debug_assets:
        urls = get_bundle_urls("web.assets_web")
        css_tags = "\n".join(f'<link rel="stylesheet" href="{u}"/>' for u in urls.get("css", []))
        js_tags = "\n".join(f'<script src="{u}"></script>' for u in urls.get("js", []))
    else:
        css_tags = '<link rel="stylesheet" href="/web/assets/web.assets_web.css"/>'
        js_tags = '<script src="/web/assets/web.assets_web.js"></script>'

    return f"""<!DOCTYPE html>
<html>
<head><meta charset="utf-8"/><meta name="viewport" content="width=device-width, initial-scale=1"/>
<title>ERP Platform</title>
{css_tags}
<style>
  body {{ margin: 0; font-family: system-ui, sans-serif; background: #f5f5f5; color: #333; min-height: 100vh; }}
  #webclient {{ min-height: 100vh; display: flex; flex-direction: column; }}
  #navbar {{ background: #1a1a2e; color: white; padding: 0.75rem 1.5rem; display: flex; align-items: center; gap: 1rem; }}
  #navbar a {{ color: white; margin-left: auto; }}
  #main {{ flex: 1; padding: 1rem; }}
</style>
</head>
<body>
<div id="webclient">
  <header id="navbar">
    <span class="logo">ERP Platform</span>
    <nav style="display:flex;gap:1rem;">
      <a href="#home" data-view="home">Home</a>
      <a href="#contacts" data-view="contacts">Contacts</a>
    </nav>
    <a href="/web/logout" style="margin-left:auto;">Logout</a>
  </header>
  <main id="main"><div id="action-manager">Loading...</div></main>
  <div id="toast-container" style="position:fixed;top:1rem;right:1rem;z-index:9999;display:flex;flex-direction:column;gap:0.5rem;pointer-events:none"></div>
</div>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.1/dist/chart.umd.min.js" crossorigin="anonymous"></script>
{js_tags}
</body>
</html>"""


@route("/web/totp/status", auth="user", methods=["GET"])
def totp_status(request):
    """Return TOTP status for current user (Phase 125)."""
    import json
    uid = get_session_uid(request)
    db = get_session_db(request)
    if uid is None:
        return Response('{"error": "unauthorized"}', status=401, content_type="application/json")
    enabled = user_has_totp_enabled(uid, db)
    return Response(json.dumps({"enabled": enabled}), content_type="application/json")


@route("/web/totp/begin_setup", auth="user", methods=["POST"])
def totp_begin_setup(request):
    """Start TOTP setup - generate secret, return provision URI (Phase 125)."""
    import json
    uid = get_session_uid(request)
    db = get_session_db(request)
    sid = request.cookies.get("erp_session")
    if uid is None or not sid:
        return Response('{"error": "unauthorized"}', status=401, content_type="application/json")
    if user_has_totp_enabled(uid, db):
        return Response('{"error": "TOTP already enabled"}', status=400, content_type="application/json")
    secret = generate_totp_secret()
    if not secret:
        return Response('{"error": "pyotp not installed. Run: pip install pyotp"}', status=500, content_type="application/json")
    with get_cursor(db) as cr:
        cr.execute("SELECT login FROM res_users WHERE id = %s", (uid,))
        row = cr.fetchone()
        login = row.get("login", "user") if row else "user"
    provision_uri = get_totp_provision_uri(secret, login)
    store_totp_setup(sid, secret, uid, db, login)
    return Response(json.dumps({"secret": secret, "provision_uri": provision_uri or ""}), content_type="application/json")


@route("/web/totp/confirm_setup", auth="user", methods=["POST"])
def totp_confirm_setup(request):
    """Confirm TOTP setup - verify code and save (Phase 125)."""
    import json
    uid = get_session_uid(request)
    db = get_session_db(request)
    sid = request.cookies.get("erp_session")
    if uid is None or not sid:
        return Response('{"error": "unauthorized"}', status=401, content_type="application/json")
    data = request.get_json(force=True, silent=True) or {}
    code = (data.get("code") or "").strip()
    if not code:
        return Response('{"error": "code required"}', status=400, content_type="application/json")
    setup = get_totp_setup(sid)
    if not setup or setup.get("uid") != uid:
        return Response('{"error": "Setup not started. Call begin_setup first."}', status=400, content_type="application/json")
    secret = setup.get("secret")
    if not verify_totp_code(uid, db, code, secret=secret):
        return Response('{"error": "Invalid code"}', status=400, content_type="application/json")
    if save_totp_to_user(uid, db, secret):
        clear_totp_setup(sid)
        return Response(json.dumps({"ok": True}), content_type="application/json")
    return Response('{"error": "Failed to save"}', status=500, content_type="application/json")


@route("/web/totp/disable", auth="user", methods=["POST"])
def totp_disable(request):
    """Disable TOTP for current user (Phase 125)."""
    import json
    uid = get_session_uid(request)
    db = get_session_db(request)
    if uid is None:
        return Response('{"error": "unauthorized"}', status=401, content_type="application/json")
    if disable_totp_for_user(uid, db):
        return Response(json.dumps({"ok": True}), content_type="application/json")
    return Response('{"error": "Failed to disable"}', status=500, content_type="application/json")


@route("/", auth="user", methods=["GET"])
def index(request):
    """Serve web client shell. Redirect to login if not authenticated. Phase 101: portal users -> /my."""
    uid = get_session_uid(request)
    if uid is None:
        return redirect("/web/login")
    db = get_session_db(request)
    try:
        from core.orm.security import get_user_groups
        registry_obj = _get_registry(db)
        groups = get_user_groups(registry_obj, db, uid)
        if "base.group_portal" in groups:
            return redirect("/my")
    except Exception:
        pass
    return Response(
        _webclient_html(debug_assets=_is_debug_assets(request)),
        content_type="text/html; charset=utf-8",
    )
